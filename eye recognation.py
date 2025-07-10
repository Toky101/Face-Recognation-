from kivy.core.window import Window
Window.clearcolor = (1, 1, 1, 1)

import cv2
import os
from datetime import datetime
from kivy.app import App
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.floatlayout import FloatLayout
from kivy.uix.image import Image
from kivy.uix.button import Button
from kivy.uix.label import Label
from kivy.clock import Clock
from kivy.graphics import Color, Line
from kivy.graphics.texture import Texture
from openpyxl import Workbook, load_workbook
import pandas as pd   
import matplotlib.pyplot as plt

# Haar cascade for face detection
face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + 'haarcascade_frontalface_default.xml')

# Setup Excel logging
base_log_dir = os.path.join(os.path.expanduser("~"), "Downloads", "EyeActivityLogs")
os.makedirs(base_log_dir, exist_ok=True)
today_folder = datetime.now().strftime("%d-%m-%Y")
current_day_output_dir = os.path.join(base_log_dir, today_folder)
os.makedirs(current_day_output_dir, exist_ok=True)
excel_path = os.path.join(current_day_output_dir, "activity_log.xlsx")
if not os.path.exists(excel_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Daily Activity Log"
    ws.append(["Start Timestamp", "End Timestamp", "Empty Duration (sec)", "Image File"])
    wb.save(excel_path)

# Helper to find connected cameras
def find_available_cameras(max_tested=5):
    available = []
    for i in range(max_tested):
        cap = cv2.VideoCapture(i)
        if cap.isOpened():
            available.append(i)
            cap.release()
    return available

class DualCameraView(BoxLayout):
    def __init__(self, **kwargs):
        super().__init__(orientation='horizontal', **kwargs)
        self.cam1_view = Image()
        self.cam2_view = Image()
        self.add_widget(self.cam1_view)
        self.add_widget(self.cam2_view)

class ThreeDButton(Button):
    def __init__(self, border_color=(0,1,0,1), text_color=(0,0.5,0,1), shadow_color=(0.7,0.7,0.7,1), **kwargs):
        super().__init__(**kwargs)
        self.background_normal = ''
        self.background_color = (1,1,1,1)
        self.color = text_color
        self.font_size = 18
        with self.canvas.before:
            Color(*shadow_color)
            self.shadow_line = Line(rounded_rectangle=(self.x+3, self.y-3, self.width, self.height, 12), width=6)
        with self.canvas.after:
            Color(*border_color)
            self.border_line = Line(rounded_rectangle=(self.x, self.y, self.width, self.height, 12), width=2)
        self.bind(pos=self.update_border, size=self.update_border)

    def update_border(self, *args):
        self.shadow_line.rounded_rectangle = (self.x+3, self.y-3, self.width, self.height, 12)
        self.border_line.rounded_rectangle = (self.x, self.y, self.width, self.height, 12)

class MainLayout(BoxLayout):
    def __init__(self, **kwargs):
        super().__init__(orientation='vertical', padding=20, spacing=10, **kwargs)

        self.video_area = DualCameraView(size_hint=(1, 0.8))
        self.add_widget(self.video_area)

        bottom_bar = FloatLayout(size_hint=(1, 0.2))
        
        self.start_button = ThreeDButton(text="Start", border_color=(0,1,0,1), text_color=(0,0.5,0,1),
                                         shadow_color=(0.7,1,0.7,1), size_hint=(0.2, 0.5), pos_hint={'x':0.08, 'y':0.2})
        self.start_button.bind(on_press=self.start_recognition)
        bottom_bar.add_widget(self.start_button)

        self.end_button = ThreeDButton(text="End", border_color=(1,0.5,0,1), text_color=(1,0.5,0,1),
                                       shadow_color=(1,0.8,0.6,1), size_hint=(0.2, 0.5), pos_hint={'right':0.92, 'y':0.2})
        self.end_button.bind(on_press=self.stop_recognition)
        bottom_bar.add_widget(self.end_button)

        self.chart_button = ThreeDButton(text="Show Chart", border_color=(0.2,0.2,1,1), text_color=(0.2,0.2,1,1),
                                         shadow_color=(0.7,0.7,1,1), size_hint=(0.2, 0.5), pos_hint={'center_x':0.5, 'y':0.2})
        self.chart_button.bind(on_press=self.show_chart)
        bottom_bar.add_widget(self.chart_button)

        self.add_widget(bottom_bar)

        self.cap1 = None
        self.cap2 = None
        self.running = False
        self.empty_start_time = None
        self.emptying = False

    def start_recognition(self, instance):
        if self.running:
            return

        available = find_available_cameras()
        print("Available cameras:", available)

        if len(available) < 1:
            print("❌ No cameras available.")
            return

        self.cap1 = cv2.VideoCapture(available[0])  # Main detection camera
        self.cap1.set(cv2.CAP_PROP_FRAME_WIDTH, 640)
        self.cap1.set(cv2.CAP_PROP_FRAME_HEIGHT, 480)

        if len(available) > 1:
            self.cap2 = cv2.VideoCapture(available[1])  # Secondary preview camera
            self.cap2.set(cv2.CAP_PROP_FRAME_WIDTH, 640)
            self.cap2.set(cv2.CAP_PROP_FRAME_HEIGHT, 480)
        else:
            self.cap2 = None
            print("⚠️ Only one camera detected.")

        self.running = True
        Clock.schedule_interval(self.process_frames, 1.0 / 30.0)

    def stop_recognition(self, instance):
        if not self.running:
            return
        self.running = False
        if self.cap1: self.cap1.release()
        if self.cap2: self.cap2.release()
        Clock.unschedule(self.process_frames)
        self.video_area.cam1_view.texture = None
        self.video_area.cam2_view.texture = None
        if self.emptying and self.empty_start_time:
            self.log_empty_period(datetime.now())

    def process_frames(self, dt):
        ret1, frame1 = self.cap1.read() if self.cap1 else (False, None)
        ret2, frame2 = self.cap2.read() if self.cap2 else (False, None)
        now = datetime.now()

        # Process main cam
        if ret1 and frame1 is not None:
            gray = cv2.cvtColor(frame1, cv2.COLOR_BGR2GRAY)
            faces = face_cascade.detectMultiScale(gray, 1.1, 5)
            person_found = len(faces) > 0

            if not person_found:
                if not self.emptying:
                    self.empty_start_time = now
                    self.emptying = True
            else:
                if self.emptying and self.empty_start_time:
                    self.log_empty_period(now, frame1)
                    self.emptying = False
                    self.empty_start_time = None

            frame1 = cv2.flip(frame1, 0)
            buf1 = frame1.tobytes()
            texture1 = Texture.create(size=(frame1.shape[1], frame1.shape[0]), colorfmt='bgr')
            texture1.blit_buffer(buf1, colorfmt='bgr', bufferfmt='ubyte')
            self.video_area.cam1_view.texture = texture1

        # Process secondary cam
        if ret2 and frame2 is not None:
            frame2 = cv2.flip(frame2, 0)
            buf2 = frame2.tobytes()
            texture2 = Texture.create(size=(frame2.shape[1], frame2.shape[0]), colorfmt='bgr')
            texture2.blit_buffer(buf2, colorfmt='bgr', bufferfmt='ubyte')
            self.video_area.cam2_view.texture = texture2

    def log_empty_period(self, end_time, frame=None):
        duration = (end_time - self.empty_start_time).total_seconds()
        start_str = self.empty_start_time.strftime("%Y-%m-%d %H:%M:%S")
        end_str = end_time.strftime("%Y-%m-%d %H:%M:%S")
        image_filename = f"empty_{end_time.strftime('%H-%M-%S')}.jpg"
        if frame is not None:
            cv2.imwrite(os.path.join(current_day_output_dir, image_filename), frame)
        wb = load_workbook(excel_path)
        ws = wb.active
        ws.append([start_str, end_str, int(duration), image_filename])
        wb.save(excel_path)
        print(f" Logged: Empty from {start_str} to {end_str} ({int(duration)}s)")

    def show_chart(self, instance):
        df = pd.read_excel(excel_path)
        total = df["Empty Duration (sec)"].sum() / 60
        plt.figure(figsize=(5,5))
        plt.bar(["Seat Empty"], [total], color='tomato')
        plt.ylabel("Minutes")
        plt.title(f"Total Time Seat Was Empty ({datetime.now().strftime('%d-%m-%Y')})")
        plt.tight_layout()
        plt.show()

class EyeRecognitionApp(App):
    def build(self):
        self.title = "Dual Camera Chair Monitor"
        return MainLayout()

    def on_stop(self):
        if hasattr(self.root, 'cap1') and self.root.cap1:
            self.root.cap1.release()
        if hasattr(self.root, 'cap2') and self.root.cap2:
            self.root.cap2.release()
        Clock.unschedule_all()

if __name__ == "__main__":
    EyeRecognitionApp().run()
